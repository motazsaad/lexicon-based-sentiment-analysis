from typing import List

from openpyxl import load_workbook
import xlsxwriter
import argparse
from io import StringIO

parser = argparse.ArgumentParser(description='Lexicon based sentiment analysis')
parser.add_argument('-pst', '--positive_file', type=argparse.FileType(mode='r', encoding='utf-8'),
                    help='lexicon for positive terms. Only one term should be per line',
                    required=True)
parser.add_argument('-neg', '--negative_file', type=argparse.FileType(mode='r', encoding='utf-8'),
                    help='lexicon for negative terms. Only one term should be per line',
                    required=True)
parser.add_argument('-f', '--filtered_tweets_file', type=str, help='xls filtered tweets.', required=True)
parser.add_argument('-opst', '--positive_tweets_file', type=str, help='xls positive tweets.', required=True)
parser.add_argument('-oneg', '--negative_tweets_file', type=str, help='xls negative tweets.', required=True)
parser.add_argument('-omix', '--mixed_tweets_file', type=str, help='xls mixed tweets.', required=True)
parser.add_argument('-oneu', '--neutral_tweets_file', type=str, help='xls neutral tweets.', required=True)


def analys_tweets(positives, negatives, filtered_tweets_work_book, positive_tweets_work_book, negative_tweets_work_book,
                  mixed_tweets_work_book, neutral_tweets_work_book):
    filtered_tweets_work_sheet = filtered_tweets_work_book['Sheet1']
    # write to positive xls
    positive_tweets_work_sheet = positive_tweets_work_book.add_worksheet()
    # negative to positive xls
    negative_tweets_work_sheet = negative_tweets_work_book.add_worksheet()
    # mixed to mixed xls
    mixed_tweets_work_sheet = mixed_tweets_work_book.add_worksheet()
    # neutral to positive xls
    neutral_tweets_work_sheet = neutral_tweets_work_book.add_worksheet()

    skip_first_row_count = 0
    pos_count = 0
    neg_count = 0
    positive_label = 0
    negative_label = 1
    mixed_label = 2
    neutral_label = 3
    label = 4
    row_positive_write = 0
    row_negative_write = 0
    row_neutral_write = 0
    row_mixed_write = 0
    col_write = 0
    for filtered_tweets_row in filtered_tweets_work_sheet.rows:
        if skip_first_row_count == 0:
            skip_first_row_count += 1
            continue
        id_value = filtered_tweets_row[0].value
        created_at = filtered_tweets_row[1].value
        text = filtered_tweets_row[2].value
        clean_text = filtered_tweets_row[3].value
        if isinstance(clean_text, str):
            for tweet_word in clean_text.split():
                tweet_word = tweet_word.strip()
                for positive_word in positives:
                    positive_word = positive_word.strip()
                    if positive_word in tweet_word:
                        pos_count += 1
                for negative_word in negatives:
                    negative_word = negative_word.strip()
                    if negative_word in tweet_word:
                        neg_count += 1
            if pos_count > neg_count:
                label = positive_label
            elif neg_count > pos_count:
                label = negative_label
            elif (pos_count == 0) and (neg_count == 0):  # neutral tweet
                label = neutral_label
            elif (pos_count == neg_count) and (pos_count > 0):  # mixed tweet, pos = negatives
                label = mixed_label
            # set counters to zero for the next tweet
            pos_count = 0
            neg_count = 0
            if label == positive_label:
                if row_positive_write == 0:
                    positive_tweets_work_sheet.write(row_positive_write, 0, 'id')
                    positive_tweets_work_sheet.write(row_positive_write, 1, 'created_at')
                    positive_tweets_work_sheet.write(row_positive_write, 2, 'text')
                    positive_tweets_work_sheet.write(row_positive_write, 3, 'clean_text')
                    row_positive_write += 1
                positive_tweets_work_sheet.write(row_positive_write, col_write, id_value)
                positive_tweets_work_sheet.write(row_positive_write, col_write + 1, created_at)
                positive_tweets_work_sheet.write(row_positive_write, col_write + 2, text)
                positive_tweets_work_sheet.write(row_positive_write, col_write + 3, clean_text)
                row_positive_write += 1
            elif label == negative_label:
                if row_negative_write == 0:
                    negative_tweets_work_sheet.write(row_negative_write, 0, 'id')
                    negative_tweets_work_sheet.write(row_negative_write, 1, 'created_at')
                    negative_tweets_work_sheet.write(row_negative_write, 2, 'text')
                    negative_tweets_work_sheet.write(row_negative_write, 3, 'clean_text')
                    row_negative_write += 1
                negative_tweets_work_sheet.write(row_negative_write, col_write, id_value)
                negative_tweets_work_sheet.write(row_negative_write, col_write + 1, created_at)
                negative_tweets_work_sheet.write(row_negative_write, col_write + 2, text)
                negative_tweets_work_sheet.write(row_negative_write, col_write + 3, clean_text)
                row_negative_write += 1
            elif label == mixed_label:
                if row_mixed_write == 0:
                    mixed_tweets_work_sheet.write(row_mixed_write, 0, 'id')
                    mixed_tweets_work_sheet.write(row_mixed_write, 1, 'created_at')
                    mixed_tweets_work_sheet.write(row_mixed_write, 2, 'text')
                    mixed_tweets_work_sheet.write(row_mixed_write, 3, 'clean_text')
                    row_mixed_write += 1
                mixed_tweets_work_sheet.write(row_mixed_write, col_write, id_value)
                mixed_tweets_work_sheet.write(row_mixed_write, col_write + 1, created_at)
                mixed_tweets_work_sheet.write(row_mixed_write, col_write + 2, text)
                mixed_tweets_work_sheet.write(row_mixed_write, col_write + 3, clean_text)
                row_mixed_write += 1
            elif label == neutral_label:
                if row_neutral_write == 0:
                    neutral_tweets_work_sheet.write(row_neutral_write, 0, 'id')
                    neutral_tweets_work_sheet.write(row_neutral_write, 1, 'created_at')
                    neutral_tweets_work_sheet.write(row_neutral_write, 2, 'text')
                    neutral_tweets_work_sheet.write(row_neutral_write, 3, 'clean_text')
                    row_neutral_write += 1
                neutral_tweets_work_sheet.write(row_neutral_write, col_write, id_value)
                neutral_tweets_work_sheet.write(row_neutral_write, col_write + 1, created_at)
                neutral_tweets_work_sheet.write(row_neutral_write, col_write + 2, text)
                neutral_tweets_work_sheet.write(row_neutral_write, col_write + 3, clean_text)
                row_neutral_write += 1
    positive_tweets_work_book.close()
    negative_tweets_work_book.close()
    mixed_tweets_work_book.close()
    neutral_tweets_work_book.close()


if __name__ == '__main__':
    args = parser.parse_args()
    positives = args.positive_file.read().splitlines()
    negatives = args.negative_file.read().splitlines()
    filtered_tweets_work_book = load_workbook(str(args.filtered_tweets_file))
    positive_tweets_work_book = xlsxwriter.Workbook(str(args.positive_tweets_file))
    negative_tweets_work_book = xlsxwriter.Workbook(str(args.negative_tweets_file))
    mixed_tweets_work_book = xlsxwriter.Workbook(str(args.mixed_tweets_file))
    neutral_tweets_work_book = xlsxwriter.Workbook(str(args.neutral_tweets_file))

    analys_tweets(positives, negatives, filtered_tweets_work_book, positive_tweets_work_book, negative_tweets_work_book,
                  mixed_tweets_work_book, neutral_tweets_work_book)


